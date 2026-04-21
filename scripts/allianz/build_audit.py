"""Build Allianz premium-rate audit workbook."""
import json
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

DATA = Path("/Users/nut_manu_td/Desktop/Financial Planner project/financial-freedom-calc/src/data/allianz/output")
OUT = Path("/Users/nut_manu_td/Desktop/Financial Planner project/financial-freedom-calc/allianz-rates-audit.xlsx")

products = json.load(open(DATA / "products.json"))["products"]
plans = json.load(open(DATA / "product_plans.json"))["plans"]
rates = json.load(open(DATA / "premium_rates.json"))["rates"]
discounts = json.load(open(DATA / "size_discounts.json"))["discounts"]
occ_mults = json.load(open(DATA / "occupation_multipliers.json"))["multipliers"]

prod_by_id = {p["id"]: p for p in products}
plan_by_id = {p["id"]: p for p in plans}

# ─── Styles ───────────────────────────────────────────────────────────────
FONT_NAME = "Arial"
TITLE_FONT = Font(name=FONT_NAME, size=14, bold=True, color="FFFFFF")
HEADER_FONT = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name=FONT_NAME, size=10)
BOLD_FONT = Font(name=FONT_NAME, size=10, bold=True)
MUTED_FONT = Font(name=FONT_NAME, size=9, italic=True, color="6B7280")
FILL_HEADER = PatternFill("solid", start_color="003781")
FILL_TITLE = PatternFill("solid", start_color="003781")
FILL_ALT = PatternFill("solid", start_color="F3F4F6")
FILL_WARN = PatternFill("solid", start_color="FEF3C7")
FILL_GOOD = PatternFill("solid", start_color="D1FAE5")
FILL_SECTION = PatternFill("solid", start_color="DBEAFE")

thin = Side(style="thin", color="D1D5DB")
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")
WRAP = Alignment(horizontal="left", vertical="center", wrap_text=True)


def set_col_widths(sheet, widths):
    for i, w in enumerate(widths, 1):
        sheet.column_dimensions[get_column_letter(i)].width = w


def apply_default_font(sheet, min_row=1):
    for row in sheet.iter_rows(min_row=min_row):
        for cell in row:
            if cell.font is None or cell.font.name != FONT_NAME:
                if not cell.font.bold:
                    cell.font = BODY_FONT


def write_title(sheet, title, width_cols):
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=width_cols)
    c = sheet.cell(row=1, column=1, value=title)
    c.font = TITLE_FONT
    c.fill = FILL_TITLE
    c.alignment = CENTER
    sheet.row_dimensions[1].height = 26


def write_headers(sheet, row, headers):
    for i, h in enumerate(headers, 1):
        cell = sheet.cell(row=row, column=i, value=h)
        cell.font = HEADER_FONT
        cell.fill = FILL_HEADER
        cell.alignment = CENTER
        cell.border = BORDER
    sheet.row_dimensions[row].height = 20


def stripe(sheet, start_row, n_rows, n_cols):
    for r in range(start_row, start_row + n_rows):
        for c in range(1, n_cols + 1):
            cell = sheet.cell(row=r, column=c)
            cell.border = BORDER
            if cell.font is None or cell.font.name != FONT_NAME:
                cell.font = BODY_FONT
            if (r - start_row) % 2 == 1:
                cell.fill = FILL_ALT


wb = Workbook()
wb.remove(wb.active)


# ═══════════════════════════════════════════════════════════════════════════
# Sheet 1: ภาพรวม (Overview)
# ═══════════════════════════════════════════════════════════════════════════
overview_data = [
    # (code, name_th, category, status, n_rates)
]
from collections import Counter
rate_count = Counter(r["product_id"] for r in rates)
for p in products:
    n = rate_count.get(p["id"], 0)
    status = "✓ พร้อม audit" if n > 0 else "⏳ ยังไม่มีข้อมูล"
    cat = "สัญญาหลัก" if p["category"] == 1 else "สัญญาเพิ่มเติม"
    overview_data.append((p["code"], p["name_th"], cat, p.get("product_type") or p.get("rider_type") or "", status, n))

# Sort: ready first, then missing
overview_data.sort(key=lambda r: (r[4] != "✓ พร้อม audit", r[0]))

ws = wb.create_sheet("ภาพรวม")
write_title(ws, "สรุปสถานะข้อมูลเบี้ย Allianz Ayudhya", 6)
ws.cell(row=2, column=1, value=f"ดึงจากตารางเบี้ยจริง • ข้อมูล ณ เม.ย. 2026 • รวม {len(products)} ผลิตภัณฑ์ ({sum(1 for r in overview_data if r[4].startswith('✓'))} มีข้อมูล, {sum(1 for r in overview_data if r[4].startswith('⏳'))} ยังไม่นำเข้า)")
ws.cell(row=2, column=1).font = MUTED_FONT
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)

write_headers(ws, 4, ["รหัส", "ชื่อผลิตภัณฑ์", "ประเภท", "กลุ่ม", "สถานะ", "จำนวน rate rows"])
for i, row in enumerate(overview_data):
    r = 5 + i
    for j, v in enumerate(row, 1):
        c = ws.cell(row=r, column=j, value=v)
        c.border = BORDER
        c.font = BODY_FONT
        if j == 1:
            c.font = BOLD_FONT
    # Colour status
    status_cell = ws.cell(row=r, column=5)
    if row[4].startswith("✓"):
        status_cell.fill = FILL_GOOD
        status_cell.font = Font(name=FONT_NAME, size=10, bold=True, color="059669")
    else:
        status_cell.fill = FILL_WARN
        status_cell.font = Font(name=FONT_NAME, size=10, color="D97706")
    ws.cell(row=r, column=6).alignment = RIGHT

set_col_widths(ws, [14, 50, 14, 14, 18, 14])
ws.freeze_panes = "A5"


# ═══════════════════════════════════════════════════════════════════════════
# Sheet 2: รายการสินค้า (Products master)
# ═══════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("สินค้าทั้งหมด")
write_title(ws, "รายการผลิตภัณฑ์ทั้งหมด (products.json)", 11)
write_headers(ws, 3, [
    "รหัส", "ชื่อไทย", "หมวด", "ชนิด/rider_type",
    "rate_per", "เพศ", "มีแผน", "มีตัวคูณอาชีพ", "มีส่วนลดทุน",
    "อายุขั้นต่ำ", "อายุสูงสุด",
])
for i, p in enumerate(sorted(products, key=lambda x: (x["category"], x["code"]))):
    r = 4 + i
    cat = "หลัก" if p["category"] == 1 else "เพิ่มเติม"
    vals = [
        p["code"], p["name_th"], cat,
        p.get("product_type") or p.get("rider_type") or "",
        p["rate_per"], p["gender_mode"],
        "ใช่" if p["has_plans"] else "—",
        "ใช่" if p["has_occ_multiplier"] else "—",
        "ใช่" if p["has_size_discount"] else "—",
        f"{p['entry_age_min']} {p['entry_age_min_unit']}",
        p["entry_age_max"],
    ]
    for j, v in enumerate(vals, 1):
        c = ws.cell(row=r, column=j, value=v)
        c.border = BORDER
        c.font = BODY_FONT
        if j in (3, 6, 7, 8, 9, 10, 11):
            c.alignment = CENTER
set_col_widths(ws, [14, 52, 10, 14, 10, 10, 9, 14, 13, 13, 12])
ws.freeze_panes = "A4"


# ═══════════════════════════════════════════════════════════════════════════
# Sheet 3: แผน (product plans)
# ═══════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("แผนผลิตภัณฑ์")
write_title(ws, "แผนย่อยของผลิตภัณฑ์ (product_plans.json)", 7)
write_headers(ws, 3, [
    "รหัสสินค้า", "ชื่อสินค้า", "รหัสแผน", "ชื่อแผน",
    "ระยะคุ้มครอง (ปี)", "ระยะจ่ายเบี้ย (ปี)", "คุ้มครองถึงอายุ",
])
for i, pl in enumerate(plans):
    r = 4 + i
    prod = prod_by_id.get(pl["product_id"], {})
    vals = [
        prod.get("code", "?"), prod.get("name_th", ""),
        pl["plan_code"], pl["plan_name_th"],
        pl.get("coverage_years") or "—",
        pl.get("premium_years") or "—",
        pl.get("coverage_until_age") or "—",
    ]
    for j, v in enumerate(vals, 1):
        c = ws.cell(row=r, column=j, value=v)
        c.border = BORDER
        c.font = BODY_FONT
        if j in (3, 5, 6, 7):
            c.alignment = CENTER
set_col_widths(ws, [14, 45, 12, 32, 16, 16, 16])
ws.freeze_panes = "A4"


# ═══════════════════════════════════════════════════════════════════════════
# Per-product rate sheets
# ═══════════════════════════════════════════════════════════════════════════
def rates_for(product_id):
    rows = [r for r in rates if r["product_id"] == product_id]
    rows.sort(key=lambda r: (r.get("plan_id") or 0, r["age_min"], r["gender"]))
    return rows


def build_rate_sheet(sheet_title, product_code, display_name, extra_note=""):
    prod = next((p for p in products if p["code"] == product_code), None)
    if not prod:
        return
    rs = rates_for(prod["id"])
    ws = wb.create_sheet(sheet_title[:31])
    n_cols = 7
    write_title(ws, f"{display_name} ({product_code})", n_cols)
    note_parts = [f"rate_per = {prod['rate_per']} บาททุน"]
    if prod.get("max_renewal_age"):
        note_parts.append(f"ต่ออายุสูงสุด {prod['max_renewal_age']} ปี")
    if prod.get("has_size_discount"):
        note_parts.append("มีส่วนลดตามทุน (ดูชีท ส่วนลดทุน)")
    if prod.get("has_occ_multiplier"):
        note_parts.append("มีตัวคูณอาชีพ (ดูชีท ค่าปรับอาชีพ)")
    if extra_note:
        note_parts.append(extra_note)
    ws.cell(row=2, column=1, value=" • ".join(note_parts)).font = MUTED_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)

    write_headers(ws, 4, [
        "แผน", "อายุต่ำสุด", "อายุสูงสุด", "เพศ",
        f"อัตราเบี้ย (บาท/{prod['rate_per']:,} ทุน)", "เฉพาะต่ออายุ?", "เบี้ยตัวอย่าง ทุน 1M",
    ])

    for i, r in enumerate(rs):
        rr = 5 + i
        plan_code = plan_by_id[r["plan_id"]]["plan_code"] if r.get("plan_id") else "—"
        gender_th = {"M": "ชาย", "F": "หญิง", "ANY": "ทุกเพศ"}.get(r["gender"], r["gender"])
        vals = [
            plan_code,
            r["age_min"], r["age_max"],
            gender_th,
            r["rate"],
            "ใช่" if r.get("is_renewal_only") else "—",
        ]
        for j, v in enumerate(vals, 1):
            c = ws.cell(row=rr, column=j, value=v)
            c.border = BORDER
            c.font = BODY_FONT
            if j in (1, 2, 3, 4, 6):
                c.alignment = CENTER
            if j == 5:
                c.alignment = RIGHT
                c.number_format = "#,##0.00"
        # Example premium formula: rate * (1,000,000 / rate_per)
        units = 1_000_000 / prod["rate_per"]
        ex = ws.cell(row=rr, column=7, value=f"=E{rr}*{units}")
        ex.border = BORDER
        ex.alignment = RIGHT
        ex.number_format = "#,##0"
        ex.font = BODY_FONT

    # Alternate-row shading
    for r in range(5, 5 + len(rs)):
        if (r - 5) % 2 == 1:
            for c in range(1, n_cols + 1):
                ws.cell(row=r, column=c).fill = FILL_ALT

    set_col_widths(ws, [10, 11, 11, 10, 22, 14, 20])
    ws.freeze_panes = "A5"


# Flat-rate products (single-row tables) — combine into one sheet for easy scan
ws = wb.create_sheet("อัตราเบี้ยแบบ Flat")
n_cols = 7
write_title(ws, "ผลิตภัณฑ์อัตราเบี้ยคงที่ (ไม่ขึ้นกับอายุ/เพศ)", n_cols)
ws.cell(row=2, column=1, value="เบี้ยต่อ 1,000 ทุน ใช้ค่าเดียวตลอดช่วงอายุ").font = MUTED_FONT
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
write_headers(ws, 4, [
    "รหัสสินค้า", "ชื่อผลิตภัณฑ์", "แผน", "ช่วงอายุ", "เพศ",
    "อัตราเบี้ย", "เบี้ยตัวอย่าง ทุน 1M",
])
flat_codes = ["MSI1808", "MDP", "MQR1206"]
row_ix = 5
for code in flat_codes:
    prod = next((p for p in products if p["code"] == code), None)
    if not prod: continue
    for r in rates_for(prod["id"]):
        plan_code = plan_by_id[r["plan_id"]]["plan_code"] if r.get("plan_id") else "—"
        gender_th = {"M": "ชาย", "F": "หญิง", "ANY": "ทุกเพศ"}.get(r["gender"], r["gender"])
        vals = [
            prod["code"], prod["name_th"], plan_code,
            f"{r['age_min']}–{r['age_max']}", gender_th, r["rate"],
        ]
        for j, v in enumerate(vals, 1):
            c = ws.cell(row=row_ix, column=j, value=v)
            c.border = BORDER
            c.font = BODY_FONT
            if j in (1, 3, 4, 5):
                c.alignment = CENTER
            if j == 6:
                c.alignment = RIGHT
                c.number_format = "#,##0.00"
        units = 1_000_000 / prod["rate_per"]
        ex = ws.cell(row=row_ix, column=7, value=f"=F{row_ix}*{units}")
        ex.border = BORDER
        ex.alignment = RIGHT
        ex.number_format = "#,##0"
        ex.font = BODY_FONT
        if (row_ix - 5) % 2 == 1:
            for c in range(1, n_cols + 1):
                ws.cell(row=row_ix, column=c).fill = FILL_ALT
        row_ix += 1
set_col_widths(ws, [14, 50, 10, 12, 10, 14, 20])
ws.freeze_panes = "A5"


# Per-product sheets with age-banded rates
build_rate_sheet("A90-21 Whole Life", "MWLA9021", "มาย โฮล ไลฟ์ A90/21")
build_rate_sheet("A99-20 Whole Life", "MWLA9920", "มาย โฮล ไลฟ์ A99/20 (มีเงินปันผล)")
build_rate_sheet("A99-6 Wealth Legacy", "MWLA9906", "มาย เวลท์ เลกาซี A99/6 (มีเงินปันผล)")
build_rate_sheet("Term 10 ปี", "T1010", "อยุธยาเฉพาะกาล 10/10")
build_rate_sheet("Term ปีต่อปี", "TM1", "อยุธยาชั่วระยะเวลา (ปีต่อปี)")
build_rate_sheet("บำนาญ A90-5", "MAFA9005", "มาย บำนาญ ไฟว์ A90/5")
build_rate_sheet("บำนาญ 85A55", "MAPA85A55", "มาย บำนาญ พลัส 85/55")
build_rate_sheet("Rider HB", "HB", "ค่ารักษาพยาบาลรายวัน (HB)", extra_note="rate คิดต่อ 100 บาทเบี้ยรายวัน")
build_rate_sheet("Rider CI48", "CI48", "โรคร้ายแรง 48 (CI48)", extra_note="rate คิดต่อ 1,000 ทุน CI")

# Tier 1 batch additions (Apr 2026)
build_rate_sheet("Rider CI48B", "CI48B", "โรคร้ายแรง 48 บียอนด์ (CI48B)",
                 extra_note="อายุ 1ด1ว–84*, occ 1–4 เบี้ยเท่ากัน")
build_rate_sheet("Rider CIMC", "CIMC", "โรคร้ายแรง มัลติ แคร์ (CIMC)",
                 extra_note="เบี้ยแบ่งตามช่วง 5 ปี, ต่ออายุถึง 98")
build_rate_sheet("Rider CB Cancer", "CB", "สัญญาคุ้มครองโรคมะเร็ง (CB)",
                 extra_note="ต่ออายุถึง 69 · คูณอาชีพ 3=1.30, 4=1.45")
build_rate_sheet("Rider CBN Cancer", "CBN", "มะเร็งหายห่วง (CBN)",
                 extra_note="3 แผน · confidence=medium ต้องตรวจซ้ำ")
build_rate_sheet("Rider TRN", "TRN", "แบบเฉพาะกาล 5/5 (TRN)",
                 extra_note="Term life rider, อายุรับ 20–59")
build_rate_sheet("Rider TRC", "TRC", "แบบเฉพาะกาล แปลงได้ (TRC)",
                 extra_note="3 แผน 10/10, 15/15, 20/20, อายุรับ 20–59")


# ═══════════════════════════════════════════════════════════════════════════
# Size discounts
# ═══════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("ส่วนลดทุน")
n_cols = 5
write_title(ws, "ส่วนลดตามขนาดทุนประกัน (size_discounts.json)", n_cols)
ws.cell(row=2, column=1, value="ส่วนลดหักจากอัตราเบี้ยต่อหน่วย เมื่อทุนอยู่ในช่วงที่กำหนด").font = MUTED_FONT
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
write_headers(ws, 4, ["รหัสสินค้า", "แผน", "ทุนขั้นต่ำ (บาท)", "ทุนสูงสุด (บาท)", "ส่วนลด (บาท/หน่วย)"])
discounts_sorted = sorted(discounts, key=lambda d: (d["product_id"], d.get("plan_id") or 0, d["sum_min"]))
for i, d in enumerate(discounts_sorted):
    r = 5 + i
    prod = prod_by_id.get(d["product_id"], {})
    plan_code = plan_by_id[d["plan_id"]]["plan_code"] if d.get("plan_id") else "—"
    vals = [prod.get("code", "?"), plan_code, d["sum_min"], d["sum_max"], d["discount_rate"]]
    for j, v in enumerate(vals, 1):
        c = ws.cell(row=r, column=j, value=v)
        c.border = BORDER
        c.font = BODY_FONT
        if j == 1 or j == 2:
            c.alignment = CENTER
        if j in (3, 4):
            c.alignment = RIGHT
            c.number_format = "#,##0"
        if j == 5:
            c.alignment = RIGHT
            c.number_format = "#,##0.00"
    if (r - 5) % 2 == 1:
        for c in range(1, n_cols + 1):
            ws.cell(row=r, column=c).fill = FILL_ALT
set_col_widths(ws, [14, 12, 20, 20, 22])
ws.freeze_panes = "A5"


# ═══════════════════════════════════════════════════════════════════════════
# Occupation multipliers
# ═══════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("ค่าปรับอาชีพ")
n_cols = 4
write_title(ws, "ตัวคูณเบี้ยตามอาชีพ (occupation_multipliers.json)", n_cols)
ws.cell(row=2, column=1, value="คูณเบี้ย rider สำหรับผู้เอาประกันตามระดับอาชีพ").font = MUTED_FONT
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=n_cols)
write_headers(ws, 4, ["รหัสสินค้า", "ชื่อสินค้า", "ระดับอาชีพ", "ตัวคูณ"])
occ_sorted = sorted(occ_mults, key=lambda o: (o["product_id"], o["occupation_class"]))
for i, m in enumerate(occ_sorted):
    r = 5 + i
    prod = prod_by_id.get(m["product_id"], {})
    vals = [prod.get("code", "?"), prod.get("name_th", ""), m["occupation_class"], m["multiplier"]]
    for j, v in enumerate(vals, 1):
        c = ws.cell(row=r, column=j, value=v)
        c.border = BORDER
        c.font = BODY_FONT
        if j in (1, 3):
            c.alignment = CENTER
        if j == 4:
            c.alignment = RIGHT
            c.number_format = "0.00"
    if (r - 5) % 2 == 1:
        for c in range(1, n_cols + 1):
            ws.cell(row=r, column=c).fill = FILL_ALT
set_col_widths(ws, [14, 40, 14, 12])
ws.freeze_panes = "A5"


# ═══════════════════════════════════════════════════════════════════════════
# Sheet: Notes for auditor
# ═══════════════════════════════════════════════════════════════════════════
ws = wb.create_sheet("หมายเหตุผู้ตรวจ")
write_title(ws, "หมายเหตุสำหรับผู้ตรวจสอบข้อมูล", 2)
notes = [
    ("วิธีอ่านไฟล์", ""),
    ("", "— ชีท 'ภาพรวม' สรุปสถานะของทั้ง 37 ผลิตภัณฑ์: ชนิดไหน audit ได้ ชนิดไหนยังไม่มีข้อมูล"),
    ("", "— ชีทต่อผลิตภัณฑ์แยกตามชื่อ (A90-21, A99-6, HB, CI48 ฯลฯ) — เทียบตารางกับ source sheet ได้โดยตรง"),
    ("", "— คอลัมน์ 'เบี้ยตัวอย่าง ทุน 1M' เป็นสูตร Excel (rate × units) เพื่อ cross-check หน้างาน"),
    ("", ""),
    ("สิ่งที่อยากให้ตรวจ", ""),
    ("", "1. อัตราเบี้ย (rate) ตรงกับตารางต้นฉบับของ Allianz หรือไม่"),
    ("", "2. ช่วงอายุ (age_min / age_max) และเพศ (ชาย / หญิง / ทุกเพศ) ตรงกันหรือไม่"),
    ("", "3. ส่วนลดทุน (ชีท 'ส่วนลดทุน') ช่วงและอัตราตรงกันหรือไม่"),
    ("", "4. ตัวคูณอาชีพ (ชีท 'ค่าปรับอาชีพ') ตรงกันหรือไม่"),
    ("", "5. ผลิตภัณฑ์ที่ขึ้น ⏳ ยังไม่มีข้อมูล ให้ระบุว่า priority สูง/กลาง/ต่ำ สำหรับการนำเข้า"),
    ("", ""),
    ("หลักเกณฑ์การคำนวณ", ""),
    ("", "สัญญาหลัก: เบี้ย = (rate − size_discount) × (ทุนประกัน ÷ rate_per)"),
    ("", "สัญญาเพิ่มเติม HB: เบี้ย = rate × (ค่ารักษา/วัน ÷ rate_per) × ตัวคูณอาชีพ"),
    ("", "สัญญาเพิ่มเติม CI: เบี้ย = rate × (ทุน ÷ rate_per) × ตัวคูณอาชีพ"),
    ("", ""),
    ("ติดต่อกลับ", ""),
    ("", "ให้แก้ในเซลล์ได้เลยแล้วส่งไฟล์กลับ พร้อม comment ระบุรายการที่พบความคลาดเคลื่อน"),
]
for i, (col1, col2) in enumerate(notes):
    r = 3 + i
    ws.cell(row=r, column=1, value=col1).font = BOLD_FONT
    ws.cell(row=r, column=1).fill = FILL_SECTION if col1 else PatternFill()
    ws.cell(row=r, column=2, value=col2).font = BODY_FONT
    ws.cell(row=r, column=2).alignment = WRAP
set_col_widths(ws, [22, 90])


wb.save(OUT)
print(f"Wrote {OUT}")
print(f"Sheets: {wb.sheetnames}")
